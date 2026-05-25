"""Export the Infra SLS schedule to two formats:

  1. CSV  (sheets-friendly, flat, one row per action)
  2. SQL  (Pumped INSERTs: wbs_nodes + actions, with placeholder org/user)

Mapping:
  Phase     -> programme        (4)
  Component -> project          (~14)
  Summary   -> work_package     (when component has summaries; else 'Default')
  Task/MS   -> action

Dependencies live in actions.notes (schema has no action-to-action FK).
"""
import csv
import re
import uuid
from datetime import datetime, timedelta
from collections import OrderedDict
from openpyxl import load_workbook

SRC = '/Users/adamhyde/Downloads/Infra SLS Schedule Task Tables — Dated.xlsx'
OUT_CSV = '/Users/adamhyde/Downloads/Infra SLS Schedule — WBS Export.csv'
OUT_SQL = '/Users/adamhyde/Downloads/Infra SLS Schedule — Pumped Import.sql'

PRED_RE = re.compile(r'^([\d\.]+)(FS|SS|FF)([+-]\d+)?d?$')

SNETS = {
    '1.1.1.1': datetime(2026, 4, 17),
    '2.1.1.1': datetime(2026, 6, 1),
    '2.2.1.1': datetime(2026, 6, 15),
    '2.3.1.1': datetime(2026, 6, 29),
    '3.1.1':   datetime(2026, 6, 15),
    '3.2.1':   datetime(2026, 11, 16),
}
TARGETS = {
    '1.1.2.3': datetime(2026, 5, 8),    '1.3.3':   datetime(2026, 5, 31),
    '2.1.5':   datetime(2026, 9, 4),    '2.2.5':   datetime(2026, 10, 9),
    '2.3.5':   datetime(2026, 11, 13),  '2.4.8':   datetime(2026, 11, 30),
    '3.1.4':   datetime(2026, 12, 21),  '3.2.2':   datetime(2027, 1, 4),
    '4.3.2':   datetime(2027, 9, 17),   '4.4.3':   datetime(2027, 11, 30),
    '4.5.9':   datetime(2027, 12, 31),
}

PHASE_NAMES = {
    'Phase 1': 'Phase 1 — Establishment',
    'Phase 2': 'Phase 2 — Capability Development',
    'Phase 3': 'Phase 3 — Build',
    'Phase 4': 'Phase 4 — Adopt and Close',
}


def workday(s, d):
    if d == 0: return s
    cur, step, rem = s, 1 if d > 0 else -1, abs(d)
    while rem:
        cur += timedelta(days=step)
        if cur.weekday() < 5: rem -= 1
    return cur


def parse_preds(s):
    if not s or not isinstance(s, str): return []
    out = []
    for tok in s.split(','):
        tok = tok.strip().replace(' ', '')
        m = PRED_RE.match(tok)
        if m:
            out.append((m.group(1), m.group(2), int(m.group(3)) if m.group(3) else 0))
    return out


def last_workday_of_month(y, m):
    nxt = datetime(y + 1, 1, 1) if m == 12 else datetime(y, m + 1, 1)
    d = nxt - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return d


# -- Recurring expansions (same as schedule model) --
def gen_monthlies(start_y, start_m, count, wbs_prefix):
    rows = []
    y, m = start_y, start_m
    for i in range(1, count + 1):
        finish = last_workday_of_month(y, m)
        start = workday(finish, -2)
        rows.append({
            'wbs': f'{wbs_prefix}.{i}',
            'task': f'Monthly report — {finish.strftime("%b %Y")}',
            'type': 'Task', 'dur': 3, 'preds': [],
            'start': start, 'finish': finish,
            'notes_extra': 'Calendar-driven',
        })
        m += 1
        if m > 12:
            m = 1
            y += 1
    return rows


RECURRING = {
    '2.6.1': {'rows': gen_monthlies(2026, 6, 6, '2.6.1'),
              'resource': 'PM + PMSO', 'component': 'Project Office', 'phase': 'Phase 2'},
    '3.3.1': {'rows': [
        {'wbs': '3.3.1.1', 'task': 'Quarterly P3M Programme Board report — Q4 2026',
         'type': 'Task', 'dur': 2, 'preds': [],
         'start': datetime(2026, 12, 29), 'finish': datetime(2026, 12, 31),
         'notes_extra': 'Calendar-driven'},
        {'wbs': '3.3.1.2', 'task': 'Quarterly P3M Programme Board report — Q1 2027',
         'type': 'Task', 'dur': 2, 'preds': [],
         'start': datetime(2027, 3, 29), 'finish': datetime(2027, 3, 31),
         'notes_extra': 'Calendar-driven'},
        {'wbs': '3.3.1.3', 'task': 'Quarterly P3M Programme Board report — Q2 2027',
         'type': 'Task', 'dur': 2, 'preds': [],
         'start': datetime(2027, 6, 28), 'finish': datetime(2027, 6, 30),
         'notes_extra': 'Calendar-driven'},
    ], 'resource': 'PM + SPO + P3MPM', 'component': 'Project Office', 'phase': 'Phase 3'},
    '3.3.2': {'rows': gen_monthlies(2027, 1, 9, '3.3.2'),
              'resource': 'PM + PMSO', 'component': 'Project Office', 'phase': 'Phase 3'},
    '4.5.6': {'rows': gen_monthlies(2027, 9, 4, '4.5.6'),
              'resource': 'PM + PMSO', 'component': 'Project Office', 'phase': 'Phase 4'},
}


# -- Load source --
ws = load_workbook(SRC, data_only=False).active

source = []
current_phase = None
current_component = None  # human-readable component name (col H text)
current_component_banner = None  # the COMP banner row (for grouping)

for row in ws.iter_rows(min_row=2, values_only=True):
    r = (list(row) + [None] * 13)[:13]
    wbs, name, typ, dur, pred, rel, res, comp, phase, start, finish, status, notes = r
    if not wbs:
        continue
    wbs = str(wbs)

    if wbs.startswith('PHASE'):
        current_phase = wbs  # e.g. 'PHASE 1' - we'll map by col I later
        source.append({'kind': 'phase_banner', 'wbs': wbs, 'name': name})
        continue
    if wbs.startswith('COMP') or wbs.startswith('GATE'):
        current_component_banner = (wbs, name)
        source.append({'kind': 'component_banner', 'wbs': wbs, 'name': name})
        continue

    source.append({
        'kind': 'row',
        'wbs': wbs,
        'task': name,
        'type': typ,
        'dur': dur,
        'pred_text': pred or '',
        'preds': parse_preds(pred or ''),
        'rel_types': rel,
        'resource': res,
        'component': comp,
        'phase': phase,
        'orig_start': start if isinstance(start, datetime) else None,
        'orig_finish': finish if isinstance(finish, datetime) else None,
        'status': status,
        'notes': notes or '',
    })


# -- Compute dates via dependency graph (same logic as verify script) --
tasks_by_wbs = {r['wbs']: r for r in source if r['kind'] == 'row'}
computed = {}

for _ in range(15):
    for wbs, t in tasks_by_wbs.items():
        if wbs in computed: continue
        if t['type'] == 'Summary': computed[wbs] = (None, None); continue
        if not isinstance(t['dur'], (int, float)):
            # recurring/calendar; will be replaced by expansion
            computed[wbs] = (None, None)
            continue
        cands = []
        if wbs in SNETS: cands.append(SNETS[wbs])
        ok = True
        for pw, pt, pl in t['preds']:
            if pw not in computed:
                ok = False; break
            ps, pf = computed[pw]
            if ps is None or pf is None:
                ok = False; break
            cands.append(workday(ps, pl) if pt == 'SS' else workday(pf, 1 + pl))
        if not ok or not cands: continue
        s = max(cands)
        f = s if t['dur'] == 0 else workday(s, int(t['dur']) - 1)
        computed[wbs] = (s, f)


# -- Build flat row list, expanding recurring --
flat_rows = []  # one per action
# Track project+wp context by scanning the source
current_phase_name = None
current_component_name = None

for entry in source:
    if entry['kind'] == 'phase_banner':
        # We resolve phase via the data row's col I, but keep banner for awareness
        continue
    if entry['kind'] == 'component_banner':
        # We resolve component via the data row's col H
        continue

    t = entry
    wbs = t['wbs']

    if wbs in RECURRING:
        # Expand
        rec = RECURRING[wbs]
        for er in rec['rows']:
            flat_rows.append({
                'wbs': er['wbs'],
                'task': er['task'],
                'type': er['type'],
                'dur': er['dur'],
                'pred_text': '',
                'preds': [],
                'rel_types': 'Calendar',
                'resource': rec['resource'],
                'component': rec['component'],
                'phase': rec['phase'],
                'start': er['start'],
                'finish': er['finish'],
                'snet': er['start'],
                'target': None,
                'notes': er.get('notes_extra', ''),
                'orig_notes': '',
            })
        continue

    if t['type'] == 'Summary':
        continue  # summaries handled implicitly via WBS hierarchy

    s, f = computed.get(wbs, (None, None))
    if t['type'] in ('Task', 'Milestone') and (s is None or f is None):
        # fallback to original typed dates
        s, f = t['orig_start'], t['orig_finish']

    flat_rows.append({
        'wbs': wbs,
        'task': t['task'],
        'type': t['type'],
        'dur': t['dur'],
        'pred_text': t['pred_text'],
        'preds': t['preds'],
        'rel_types': t['rel_types'],
        'resource': t['resource'],
        'component': t['component'],
        'phase': t['phase'],
        'start': s,
        'finish': f,
        'snet': SNETS.get(wbs),
        'target': TARGETS.get(wbs),
        'notes': t['notes'],
        'orig_notes': t['notes'],
    })


# -- Derive WBS tree: programme (phase) -> project (component) -> work_package (summary) -> action --
def summary_wbs_of(action_wbs):
    """For action '1.1.1.5' return summary '1.1.1'; for '1.3.1' return None
    (no summary parent — direct under component)."""
    parts = action_wbs.split('.')
    if len(parts) >= 4:
        return '.'.join(parts[:3])
    return None


# Component lookup: each task row has col H (component_name) and col I (phase)
# We need WP groupings under each component. Use the summary WBS code if it exists
# in source (as a Summary row), otherwise the component gets a single 'Default' WP.

summary_names = {}
for r in source:
    if r['kind'] != 'row': continue
    if r['type'] == 'Summary':
        summary_names[r['wbs']] = r['task']

# Build programme/project/wp tables
programmes = OrderedDict()   # phase_key -> {name, id}
projects = OrderedDict()     # (phase_key, component_name) -> {name, parent: phase_key, id}
work_packages = OrderedDict()  # (phase_key, component, wp_key) -> {name, parent: project_key, id}

for r in flat_rows:
    phase_key = r['phase']
    comp = r['component'] or 'Unassigned'
    proj_key = (phase_key, comp)
    if phase_key not in programmes:
        programmes[phase_key] = {'name': PHASE_NAMES.get(phase_key, phase_key), 'id': str(uuid.uuid4())}
    if proj_key not in projects:
        projects[proj_key] = {'name': comp, 'parent': phase_key, 'id': str(uuid.uuid4())}
    # WP
    summary_wbs = summary_wbs_of(r['wbs'])
    if summary_wbs and summary_wbs in summary_names:
        wp_key = (phase_key, comp, summary_wbs)
        wp_name = f"{summary_wbs} {summary_names[summary_wbs]}"
    else:
        wp_key = (phase_key, comp, 'default')
        wp_name = f"{comp} — Tasks"
    if wp_key not in work_packages:
        work_packages[wp_key] = {'name': wp_name, 'parent': proj_key, 'id': str(uuid.uuid4())}
    r['_wp_key'] = wp_key


# -- Write CSV --
with open(OUT_CSV, 'w', newline='') as fh:
    w = csv.writer(fh)
    w.writerow([
        'WBS', 'Programme', 'Project', 'Work Package', 'Task', 'Type',
        'Duration (wd)', 'Start Date', 'Finish Date', 'SNET', 'Target Date',
        'Predecessors', 'Resources', 'Notes',
    ])
    for r in flat_rows:
        wp = work_packages[r['_wp_key']]['name']
        proj = projects[(r['phase'], r['component'] or 'Unassigned')]['name']
        prog = programmes[r['phase']]['name']
        w.writerow([
            r['wbs'], prog, proj, wp, r['task'], r['type'],
            r['dur'] if isinstance(r['dur'], (int, float)) else (r['dur'] or ''),
            r['start'].strftime('%Y-%m-%d') if r['start'] else '',
            r['finish'].strftime('%Y-%m-%d') if r['finish'] else '',
            r['snet'].strftime('%Y-%m-%d') if r['snet'] else '',
            r['target'].strftime('%Y-%m-%d') if r['target'] else '',
            r['pred_text'], r['resource'] or '', r['notes'] or '',
        ])

print(f"CSV: {OUT_CSV}  ({len(flat_rows)} actions)")


# -- Build SQL --
def sql_escape(s):
    if s is None: return 'NULL'
    return "'" + str(s).replace("'", "''") + "'"


def sql_date(d):
    return f"'{d.strftime('%Y-%m-%d')}'" if d else 'NULL'


def labels_for(r):
    labels = [r['type'].lower(), r['phase'].lower().replace(' ', '-')]
    if r['resource']:
        # split resources roughly on '+' or ',' for tags
        for tok in re.split(r'[+,]', r['resource']):
            tok = tok.strip()
            if tok and len(tok) <= 20:
                labels.append('res:' + tok.lower())
    return labels


def sql_array(items):
    if not items: return "ARRAY[]::text[]"
    return "ARRAY[" + ", ".join(sql_escape(x) for x in items) + "]"


def make_notes(r):
    parts = []
    if r['pred_text']:
        parts.append(f"Predecessors: {r['pred_text']}")
    if r['resource']:
        parts.append(f"Resources: {r['resource']}")
    if isinstance(r['dur'], (int, float)):
        parts.append(f"Duration: {int(r['dur'])} working days")
    if r['snet']:
        parts.append(f"SNET: {r['snet'].strftime('%d %b %Y')}")
    if r['target']:
        parts.append(f"Gate target: {r['target'].strftime('%d %b %Y')}")
    if r['notes']:
        parts.append(r['notes'])
    parts.append(f"WBS: {r['wbs']}")
    return '\n'.join(parts)


lines = []
lines.append("-- Pumped bulk import: Infra SLS Schedule")
lines.append("-- Replace :org_id and :user_id below before running, or set as psql variables.")
lines.append("-- Run as a single transaction in Supabase SQL Editor.")
lines.append("")
lines.append("BEGIN;")
lines.append("")
lines.append("-- ============= Replace these two values =============")
lines.append("\\set org_id  '00000000-0000-0000-0000-000000000000'")
lines.append("\\set user_id '00000000-0000-0000-0000-000000000000'")
lines.append("-- =====================================================")
lines.append("")

# Programmes
lines.append("-- Programmes (one per phase)")
lines.append(
    "INSERT INTO wbs_nodes (id, organisation_id, parent_id, node_type, name, "
    "description, position, created_by) VALUES"
)
prog_rows = []
for pos, (key, p) in enumerate(programmes.items()):
    prog_rows.append(
        f"  ('{p['id']}', :'org_id', NULL, 'programme', "
        f"{sql_escape(p['name'])}, '', {pos}, :'user_id')"
    )
lines.append(',\n'.join(prog_rows) + ';')
lines.append('')

# Projects
lines.append("-- Projects (one per component)")
lines.append(
    "INSERT INTO wbs_nodes (id, organisation_id, parent_id, node_type, name, "
    "description, position, project_status, created_by) VALUES"
)
proj_rows = []
for pos, (key, p) in enumerate(projects.items()):
    parent_id = programmes[p['parent']]['id']
    proj_rows.append(
        f"  ('{p['id']}', :'org_id', '{parent_id}', 'project', "
        f"{sql_escape(p['name'])}, '', {pos}, 'active', :'user_id')"
    )
lines.append(',\n'.join(proj_rows) + ';')
lines.append('')

# Work packages
lines.append("-- Work Packages (one per summary group, or default per project)")
lines.append(
    "INSERT INTO wbs_nodes (id, organisation_id, parent_id, node_type, name, "
    "description, position, rag_status, created_by) VALUES"
)
wp_rows = []
for pos, (key, w) in enumerate(work_packages.items()):
    parent_id = projects[w['parent']]['id']
    wp_rows.append(
        f"  ('{w['id']}', :'org_id', '{parent_id}', 'work_package', "
        f"{sql_escape(w['name'])}, '', {pos}, 'green', :'user_id')"
    )
lines.append(',\n'.join(wp_rows) + ';')
lines.append('')

# Actions
lines.append("-- Actions (one per task / milestone)")
lines.append(
    "INSERT INTO actions (id, organisation_id, wbs_node_id, task, priority, "
    "status, start_date, due_date, notes, labels, created_by) VALUES"
)
act_rows = []
for pos, r in enumerate(flat_rows):
    wp_id = work_packages[r['_wp_key']]['id']
    aid = str(uuid.uuid4())
    priority = 'high' if r['type'] == 'Milestone' else 'medium'
    act_rows.append(
        f"  ('{aid}', :'org_id', '{wp_id}', {sql_escape(r['task'])}, "
        f"'{priority}', 'not_started', {sql_date(r['start'])}, {sql_date(r['finish'])}, "
        f"{sql_escape(make_notes(r))}, {sql_array(labels_for(r))}, :'user_id')"
    )
lines.append(',\n'.join(act_rows) + ';')
lines.append('')

lines.append("COMMIT;")
lines.append('')
lines.append(f"-- Summary: {len(programmes)} programmes, {len(projects)} projects, "
             f"{len(work_packages)} work packages, {len(flat_rows)} actions.")

with open(OUT_SQL, 'w') as fh:
    fh.write('\n'.join(lines))

print(f"SQL: {OUT_SQL}")
print(f"  {len(programmes)} programmes, {len(projects)} projects, "
      f"{len(work_packages)} work packages, {len(flat_rows)} actions")
