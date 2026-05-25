"""Simulate the formula network in Python and compare to the original
hand-typed dates. Any divergence is either a formula bug on my side or a
hand-arithmetic slip in the source sheet."""
import re
from datetime import datetime, timedelta
from openpyxl import load_workbook

SRC = '/Users/adamhyde/Downloads/Infra SLS Schedule Task Tables — Dated.xlsx'

PRED_RE = re.compile(r'^([\d\.]+)(FS|SS|FF)([+-]\d+)?d?$')

SNETS = {
    '1.1.1.1': datetime(2026, 4, 17),
    '2.1.1.1': datetime(2026, 6, 1),
    '2.2.1.1': datetime(2026, 6, 15),
    '2.3.1.1': datetime(2026, 6, 29),
    '3.1.1':   datetime(2026, 6, 15),
    '3.2.1':   datetime(2026, 11, 16),
}


def workday(start, days):
    if days == 0:
        return start
    d = start
    step = 1 if days > 0 else -1
    remaining = abs(days)
    while remaining > 0:
        d += timedelta(days=step)
        if d.weekday() < 5:
            remaining -= 1
    return d


def parse_preds(s):
    if not s or not isinstance(s, str):
        return []
    out = []
    for tok in s.split(','):
        tok = tok.strip().replace(' ', '')
        if not tok:
            continue
        m = PRED_RE.match(tok)
        if m:
            wbs, rtype, lag = m.group(1), m.group(2), m.group(3)
            out.append((wbs, rtype, int(lag) if lag else 0))
    return out


wb = load_workbook(SRC, data_only=False)
ws = wb.active

tasks = {}
order = []
for row in ws.iter_rows(min_row=2, values_only=True):
    wbs, name, typ, dur, pred, rel, res, comp, phase, start, finish, status, notes = (
        list(row) + [None] * (13 - len(row))
    )[:13]
    if not wbs:
        continue
    wbs = str(wbs)
    if typ not in ('Task', 'Milestone', 'Summary'):
        continue
    tasks[wbs] = {
        'name': name,
        'type': typ,
        'dur': dur if isinstance(dur, (int, float)) else None,
        'pred_text': pred,
        'orig_start': start if isinstance(start, datetime) else None,
        'orig_finish': finish if isinstance(finish, datetime) else None,
        'preds': parse_preds(pred or ''),
    }
    order.append(wbs)


# Compute start/finish in topological-ish order. Iterate until stable.
computed = {}

def compute_start(wbs):
    t = tasks[wbs]
    candidates = []
    snet = SNETS.get(wbs)
    if snet:
        candidates.append(snet)
    for pw, ptype, plag in t['preds']:
        if pw not in computed:
            return None  # predecessor not ready
        ps, pf = computed[pw]
        if ps is None or pf is None:
            return None
        if ptype == 'SS':
            candidates.append(workday(ps, plag))
        else:  # FS (or FF treated as FS for now)
            candidates.append(workday(pf, 1 + plag))
    if not candidates:
        return None
    return max(candidates)


def compute_finish(start, dur):
    if start is None or dur is None:
        return None
    if dur == 0:
        return start
    return workday(start, int(dur) - 1)


# Iterate
for _ in range(10):
    progress = False
    for wbs in order:
        if wbs in computed:
            continue
        t = tasks[wbs]
        if t['type'] == 'Summary':
            computed[wbs] = (None, None)
            progress = True
            continue
        s = compute_start(wbs)
        if s is None:
            continue
        f = compute_finish(s, t['dur'])
        computed[wbs] = (s, f)
        progress = True
    if not progress:
        break

# Compare
mismatches = []
for wbs in order:
    t = tasks[wbs]
    if t['type'] == 'Summary':
        continue
    if wbs not in computed:
        mismatches.append((wbs, 'NOT COMPUTED', t['name']))
        continue
    s, f = computed[wbs]
    os_, of_ = t['orig_start'], t['orig_finish']
    if os_ and s and s.date() != os_.date():
        mismatches.append((wbs, f'start: computed={s.date()} typed={os_.date()}', t['name']))
    if of_ and f and f.date() != of_.date():
        mismatches.append((wbs, f'finish: computed={f.date()} typed={of_.date()}', t['name']))

total = len([w for w in order if tasks[w]['type'] != 'Summary'])
print(f"Compared {total} task/milestone rows.")
print(f"Matches: {total - len(mismatches)}  Mismatches: {len(mismatches)}")
print()
for m in mismatches:
    print(f"  {m[0]:12}  {m[1]}")
    print(f"               {m[2]}")
