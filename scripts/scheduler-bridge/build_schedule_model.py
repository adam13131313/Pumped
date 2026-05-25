"""Rebuild the Infra SLS schedule as a formula-driven model.

Conventions verified against the source data:
  FS+X  ->  successor.start = WORKDAY(pred.finish, 1+X)
  SS+X  ->  successor.start = WORKDAY(pred.start, X)
  Finish = WORKDAY(start, dur-1)   for dur >= 1
  Finish = Start                    for dur = 0 (milestones)
  No holidays calendar -- M-F only.

Helpers (Q..AF) hold parsed predecessor tokens; Schedule (J..N) holds
SNET/Start/Finish/Target/Variance with formulas chained via INDEX/MATCH
on the WBS column.
"""
import re
from datetime import datetime, timedelta, date
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

SRC = '/Users/adamhyde/Downloads/Infra SLS Schedule Task Tables — Dated.xlsx'
OUT = '/Users/adamhyde/Downloads/Infra SLS Schedule — Formula Driven.xlsx'

PRED_RE = re.compile(r'^([\d\.]+)(FS|SS|FF)([+-]\d+)?d?$')


def parse_preds(pred_str):
    if not pred_str or not isinstance(pred_str, str):
        return []
    out = []
    for tok in pred_str.split(','):
        tok = tok.strip().replace(' ', '')
        if not tok:
            continue
        m = PRED_RE.match(tok)
        if m:
            wbs, rtype, lag = m.group(1), m.group(2), m.group(3)
            out.append((wbs, rtype, int(lag) if lag else 0))
        else:
            print(f"  ! unparseable predecessor token: {tok!r}")
    return out


def last_workday_of_month(y, m):
    if m == 12:
        nxt = date(y + 1, 1, 1)
    else:
        nxt = date(y, m + 1, 1)
    d = nxt - timedelta(days=1)
    while d.weekday() >= 5:
        d -= timedelta(days=1)
    return datetime(d.year, d.month, d.day)


# --- Extracted from the original notes column ---
SNETS = {
    '1.1.1.1': datetime(2026, 4, 17),   # project anchor
    '2.1.1.1': datetime(2026, 6, 1),
    '2.2.1.1': datetime(2026, 6, 15),
    '2.3.1.1': datetime(2026, 6, 29),
    '3.1.1':   datetime(2026, 6, 15),
    '3.2.1':   datetime(2026, 11, 16),
}

TARGETS = {
    '1.1.2.3': datetime(2026, 5, 8),
    '1.3.3':   datetime(2026, 5, 31),
    '2.1.5':   datetime(2026, 9, 4),
    '2.2.5':   datetime(2026, 10, 9),
    '2.3.5':   datetime(2026, 11, 13),
    '2.4.8':   datetime(2026, 11, 30),
    '3.1.4':   datetime(2026, 12, 21),
    '3.2.2':   datetime(2027, 1, 4),
    '4.3.2':   datetime(2027, 9, 17),
    '4.4.3':   datetime(2027, 11, 30),
    '4.5.9':   datetime(2027, 12, 31),
}


# --- Recurring rows to expand ---
def gen_monthlies(start_y, start_m, count, wbs_prefix, base_name):
    rows = []
    y, m = start_y, start_m
    for i in range(1, count + 1):
        finish = last_workday_of_month(y, m)
        rows.append({
            'wbs': f'{wbs_prefix}.{i}',
            'task': f'{base_name} — {finish.strftime("%b %Y")}',
            'type': 'Task',
            'dur': 3,
            'pred_text': '',
            'rel_types': 'Calendar',
            'snet': finish,  # treat as calendar anchor
            'is_recurring': True,
        })
        m += 1
        if m > 12:
            m = 1
            y += 1
    return rows


RECURRING_EXPANSIONS = {
    # Original WBS -> list of expanded rows
    '2.6.1': {
        'rows': gen_monthlies(2026, 6, 6, '2.6.1', 'Monthly report'),
        'resource': 'PM + PMSO',
        'component': 'Project Office',
        'phase': 'Phase 2',
    },
    '3.3.1': {
        'rows': [
            {'wbs': '3.3.1.1', 'task': 'Quarterly P3M Programme Board report — Q4 2026',
             'type': 'Task', 'dur': 2, 'pred_text': '', 'rel_types': 'Calendar',
             'snet': datetime(2026, 12, 31), 'is_recurring': True},
            {'wbs': '3.3.1.2', 'task': 'Quarterly P3M Programme Board report — Q1 2027',
             'type': 'Task', 'dur': 2, 'pred_text': '', 'rel_types': 'Calendar',
             'snet': datetime(2027, 3, 31), 'is_recurring': True},
            {'wbs': '3.3.1.3', 'task': 'Quarterly P3M Programme Board report — Q2 2027',
             'type': 'Task', 'dur': 2, 'pred_text': '', 'rel_types': 'Calendar',
             'snet': datetime(2027, 6, 30), 'is_recurring': True},
        ],
        'resource': 'PM + SPO + P3MPM',
        'component': 'Project Office',
        'phase': 'Phase 3',
    },
    '3.3.2': {
        'rows': gen_monthlies(2027, 1, 9, '3.3.2', 'Monthly report'),
        'resource': 'PM + PMSO',
        'component': 'Project Office',
        'phase': 'Phase 3',
    },
    '4.5.6': {
        'rows': gen_monthlies(2027, 9, 4, '4.5.6', 'Monthly report'),
        'resource': 'PM + PMSO',
        'component': 'Project Office',
        'phase': 'Phase 4',
    },
}


# --- Load source ---
src = load_workbook(SRC, data_only=False)
src_ws = src.active

src_rows = []
for row in src_ws.iter_rows(min_row=2, max_row=src_ws.max_row, values_only=True):
    src_rows.append(list(row) + [None] * (13 - len(row)))


# Convert source row to internal task dict
def row_to_task(r):
    wbs, task, typ, dur, pred, rel, res, comp, phase, start, finish, status, notes = r
    return {
        'wbs': str(wbs) if wbs is not None else '',
        'task': task,
        'type': typ,
        'dur': dur,
        'pred_text': pred or '',
        'rel_types': rel,
        'resource': res,
        'component': comp,
        'phase': phase,
        'orig_start': start,
        'orig_finish': finish,
        'status': status,
        'notes': notes,
    }


# Build output row list (expanding recurring)
out_rows = []
for r in src_rows:
    t = row_to_task(r)
    if t['wbs'] in RECURRING_EXPANSIONS:
        exp = RECURRING_EXPANSIONS[t['wbs']]
        # Insert a summary banner using the original WBS row
        for er in exp['rows']:
            out_rows.append({
                **er,
                'resource': exp['resource'],
                'component': exp['component'],
                'phase': exp['phase'],
                'notes': 'Calendar-driven',
            })
    else:
        out_rows.append(t)


# --- Build output workbook ---
out = Workbook()
ws = out.active
ws.title = 'Schedule'

HEADERS = [
    'WBS', 'Task', 'Type', 'Dur (d)', 'Predecessor(s)', 'Rel Types',
    'Resource(s)', 'Component', 'Phase',
    'SNET', 'Start Date', 'Finish Date', 'Target Date', 'Variance (wd)',
    'Schedule Status', 'Notes',
    # Helpers
    'P1_WBS', 'P1_Type', 'P1_Lag', 'P1_Start',
    'P2_WBS', 'P2_Type', 'P2_Lag', 'P2_Start',
    'P3_WBS', 'P3_Type', 'P3_Lag', 'P3_Start',
    'P4_WBS', 'P4_Type', 'P4_Lag', 'P4_Start',
]

for ci, h in enumerate(HEADERS, start=1):
    c = ws.cell(row=1, column=ci, value=h)
    c.font = Font(bold=True)
    c.alignment = Alignment(horizontal='center', vertical='center')

# Excel column letters
COL = {h: get_column_letter(i + 1) for i, h in enumerate(HEADERS)}

# Write data rows
DATA_START = 2
for ri, t in enumerate(out_rows):
    row = DATA_START + ri
    ws.cell(row=row, column=1, value=t.get('wbs'))
    ws.cell(row=row, column=2, value=t.get('task'))
    ws.cell(row=row, column=3, value=t.get('type'))
    dur = t.get('dur')
    # Handle non-numeric durations from recurring originals (we already expanded them)
    ws.cell(row=row, column=4, value=dur if isinstance(dur, (int, float)) else dur)
    ws.cell(row=row, column=5, value=t.get('pred_text'))
    ws.cell(row=row, column=6, value=t.get('rel_types'))
    ws.cell(row=row, column=7, value=t.get('resource'))
    ws.cell(row=row, column=8, value=t.get('component'))
    ws.cell(row=row, column=9, value=t.get('phase'))

    wbs = t.get('wbs', '')
    snet = t.get('snet') or SNETS.get(wbs)
    if snet is not None:
        ws.cell(row=row, column=10, value=snet).number_format = 'd-mmm-yyyy'

    # Target
    tgt = TARGETS.get(wbs)
    if tgt is not None:
        ws.cell(row=row, column=13, value=tgt).number_format = 'd-mmm-yyyy'

    # Notes
    ws.cell(row=row, column=16, value=t.get('notes'))

    # Parse predecessors into helper columns
    preds = parse_preds(t.get('pred_text', ''))
    for pi, (pw, ptype, plag) in enumerate(preds[:4]):
        base_col = 17 + pi * 4  # Q=17 for P1, U=21 for P2, Y=25, AC=29
        ws.cell(row=row, column=base_col, value=pw)
        ws.cell(row=row, column=base_col + 1, value=ptype)
        ws.cell(row=row, column=base_col + 2, value=plag)


# Pass 2: write formulas. Need INDEX/MATCH on WBS column.
WBS_RANGE = f"$A$2:$A${DATA_START + len(out_rows) - 1}"
START_RANGE = f"$K$2:$K${DATA_START + len(out_rows) - 1}"
FINISH_RANGE = f"$L$2:$L${DATA_START + len(out_rows) - 1}"

# helper to make a candidate-start formula for a predecessor
def cand_start_formula(row, pred_n):
    """Formula for P{n}_Start cell. Looks up pred row by WBS, picks
    Start (if SS) or Finish (if FS) of predecessor, applies lag, returns date."""
    base_col = 17 + (pred_n - 1) * 4
    wbs_cell = f"{get_column_letter(base_col)}{row}"
    type_cell = f"{get_column_letter(base_col + 1)}{row}"
    lag_cell = f"{get_column_letter(base_col + 2)}{row}"
    # If no WBS, return blank
    return (
        f'=IF({wbs_cell}="","",'
        f'IF({type_cell}="SS",'
        f'WORKDAY(INDEX({START_RANGE},MATCH({wbs_cell},{WBS_RANGE},0)),{lag_cell}),'
        f'WORKDAY(INDEX({FINISH_RANGE},MATCH({wbs_cell},{WBS_RANGE},0)),1+{lag_cell})'
        f'))'
    )


for ri in range(len(out_rows)):
    row = DATA_START + ri
    t = out_rows[ri]

    # P1..P4 candidate-start formulas
    for n in range(1, 5):
        col = 17 + (n - 1) * 4 + 3  # T=20 (P1_Start), X=24 (P2_Start), AB=28, AF=32
        ws.cell(row=row, column=col, value=cand_start_formula(row, n))
        ws.cell(row=row, column=col).number_format = 'd-mmm-yyyy'

    # Start (col K=11): MAX(SNET, all candidate starts). If summary row, leave blank.
    typ = t.get('type')
    if typ == 'Summary' or typ is None:
        # Don't compute for summary banners
        continue

    p_starts = [f"{get_column_letter(17 + (n-1)*4 + 3)}{row}" for n in range(1, 5)]
    snet_cell = f"$J${row}"
    # MAX treats "" as 0, so we need IFERROR / IF to handle blanks
    # We'll use MAX with the SNET and the candidate starts; blanks become 0,
    # MAX picks the largest real date.
    start_formula = (
        f'=IFERROR(MAX('
        f'IF({snet_cell}="",0,{snet_cell}),'
        + ",".join(f'IF({c}="",0,{c})' for c in p_starts)
        + '),"")'
    )
    start_cell = ws.cell(row=row, column=11, value=start_formula)
    start_cell.number_format = 'd-mmm-yyyy'

    # Finish (col L=12): WORKDAY(Start, Dur-1) for dur>=1; Start for dur=0
    start_ref = f"K{row}"
    dur_ref = f"D{row}"
    finish_formula = (
        f'=IF({start_ref}="","",'
        f'IF({dur_ref}=0,{start_ref},WORKDAY({start_ref},{dur_ref}-1)))'
    )
    finish_cell = ws.cell(row=row, column=12, value=finish_formula)
    finish_cell.number_format = 'd-mmm-yyyy'

    # Variance vs target (col N=14): only if target exists
    if t.get('wbs') in TARGETS:
        # Variance = working days between Target and Finish (positive = late)
        # NETWORKDAYS(target, finish) - 1, with sign
        var_formula = (
            f'=IF(OR(L{row}="",M{row}=""),"",'
            f'IF(L{row}>=M{row},NETWORKDAYS(M{row},L{row})-1,'
            f'-(NETWORKDAYS(L{row},M{row})-1)))'
        )
        ws.cell(row=row, column=14, value=var_formula)

# Formatting: phase/component/gate banner rows
banner_fill = PatternFill('solid', start_color='D9E1F2')
component_fill = PatternFill('solid', start_color='E7E6E6')
gate_fill = PatternFill('solid', start_color='FFE699')
milestone_fill = PatternFill('solid', start_color='FFF2CC')
header_fill = PatternFill('solid', start_color='305496')

for ci in range(1, len(HEADERS) + 1):
    c = ws.cell(row=1, column=ci)
    c.fill = header_fill
    c.font = Font(bold=True, color='FFFFFF')

for ri, t in enumerate(out_rows):
    row = DATA_START + ri
    wbs = t.get('wbs', '') or ''
    typ = t.get('type')
    fill = None
    bold = False
    if wbs.startswith('PHASE'):
        fill = banner_fill
        bold = True
    elif wbs.startswith('COMP') or wbs.startswith('GATE'):
        fill = component_fill
        bold = True
    elif typ == 'Milestone':
        # Gate milestones get gold; regular milestones get pale yellow
        is_gate = wbs in TARGETS or 'GATE' in str(t.get('notes') or '')
        fill = gate_fill if is_gate else milestone_fill
        bold = is_gate
    elif typ == 'Summary':
        fill = component_fill

    if fill or bold:
        for ci in range(1, 17):  # first 16 cols (model)
            c = ws.cell(row=row, column=ci)
            if fill:
                c.fill = fill
            if bold:
                c.font = Font(bold=True)

# Column widths
widths = {
    'A': 9, 'B': 56, 'C': 10, 'D': 8, 'E': 28, 'F': 12, 'G': 22,
    'H': 22, 'I': 9, 'J': 12, 'K': 12, 'L': 12, 'M': 12, 'N': 10,
    'O': 14, 'P': 38,
}
for col, w in widths.items():
    ws.column_dimensions[col].width = w
# Helper columns narrow
for col_idx in range(17, 33):
    ws.column_dimensions[get_column_letter(col_idx)].width = 8

# Freeze headers
ws.freeze_panes = 'C2'

# Group helper columns so they collapse
ws.column_dimensions.group('Q', 'AF', hidden=False, outline_level=1)

# Helper sheet header styling
for ci in range(17, 33):
    c = ws.cell(row=1, column=ci)
    c.fill = PatternFill('solid', start_color='8497B0')
    c.font = Font(bold=True, color='FFFFFF', size=9)

# --- Anchor / explanatory notes at top? Skip; the SNET column on row 2 (1.1.1.1) IS the anchor.

# Save
out.save(OUT)
print(f"Wrote {OUT}")
print(f"Rows: {len(out_rows)} (data) + 1 (header)")
